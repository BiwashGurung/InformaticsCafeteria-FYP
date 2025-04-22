from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth import authenticate, login , logout
from django.contrib.auth.decorators import user_passes_test
from django.contrib import messages
from .models import EventPopup
from .forms import EventPopupForm, FoodItemForm 
from datetime import datetime , timedelta
from django.utils import timezone
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from django import forms
from django.contrib.auth.models import User
from django.contrib.auth.hashers import make_password
# Importing Profile from cafeteria app
from cafeteria.models import Profile  , FoodItem , OrderItem, Order , LostFound , GroupOrder, Feedback , Reply
from django.db.models import Q
from django.db.models import Sum
import logging


logger = logging.getLogger(__name__)
#Helper function to check whether the user is an admin or not
def is_admin(user):
    return user.is_staff

def cafeteria_admin_login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            if user.is_staff:
                login(request, user)
                return redirect('/cafeteria_admin/dashboard/')
            else:
                messages.error(request, 'You are not authorized to access admin panel.')
        else:
            messages.error(request, 'Invalid admin username or password.')
    return render(request, 'cafeteria_admin/admin_login.html')


# Admin dashboard view
@user_passes_test(is_admin, login_url='/cafeteria_admin/admin_login/')
def cafeteria_admin_dashboard(request):
    # Counting the total number of users
    total_users = Profile.objects.count()  
    total_orders = Order.objects.count()
    order_items = OrderItem.objects.count()
    pending_orders = Order.objects.filter(status='Pending').select_related('user').order_by('-order_date')[:5]
    total_revenue = OrderItem.objects.aggregate(total=Sum('price'))['total'] or 0

    context = {
        'total_users': total_users,
        'total_orders': total_orders,  
        'order_items': order_items,
        'pending_orders': pending_orders,
        'total_revenue': total_revenue, 
    }

    return render(request, 'cafeteria_admin/dashboard.html', context)
#Logout View
@user_passes_test(is_admin, login_url='/cafeteria_admin/admin_login/')
def logout_admin(request):
    logout(request)  
    return redirect('/cafeteria_admin/admin_login/')


#Allowing admin to upload events
@user_passes_test(is_admin, login_url='/cafeteria_admin/admin_login/')
def admin_upload_popup(request):
    if request.method == 'POST':
        #Processing  the form data
        form = EventPopupForm(request.POST, request.FILES)
        if form.is_valid():
            #Saving the form if valid
            form.save()
            #Redirecting after successful submisson of event
            return redirect('admin_upload_popup')  
    else:
        #Creating an empty form for GET requests
        form = EventPopupForm()

    return render(request, 'cafeteria_admin/event_popup.html', {'form': form})

#Displayinh tthe curremt popup event to the users (homepage)
def show_popup(request):
    #Getting the current data and time
    current_time = datetime.now()
    #Getting  the latest event within the current time range
    event = EventPopup.objects.filter(start_date__lte=current_time, end_date__gte=current_time).order_by('-start_date').first()  
    #Limiting the number of food items to be displayed on the homepage
    food_items = FoodItem.objects.all()[:4] 
    #Passing the even to the homepage
    return render(request, 'cafeteria/index.html', {'event': event, 'food_items': food_items})

@user_passes_test(is_admin, login_url='/cafeteria_admin/admin_login/')
def view_event_history(request):
    # Fetching all events ordered by newest first
    events = EventPopup.objects.all().order_by('-start_date')  
    return render(request, 'cafeteria_admin/view_event_history.html', {'events': events})


@user_passes_test(is_admin, login_url='/cafeteria_admin/admin_login/')
def edit_event(request, event_id):
    try:
        event = get_object_or_404(EventPopup, event_id=event_id)
        if request.method == 'POST':
            form = EventPopupForm(request.POST, request.FILES, instance=event)
            if form.is_valid():
                form.save()
                messages.success(request, f"Event '{event.event_title}' updated successfully!")
                return redirect('view_event_history')
            else:
                messages.error(request, 'Error updating event. Please check the form.')
        else:
            form = EventPopupForm(instance=event)
        return render(request, 'cafeteria_admin/edit_event.html', {'form': form, 'event': event})
    except Exception as e:
        logger.error(f"Error in edit_event: {str(e)}")
        messages.error(request, "An error occurred while editing the event.")
        return redirect('view_event_history')

@user_passes_test(is_admin, login_url='/cafeteria_admin/admin_login/')
def delete_event(request, event_id):
    event = get_object_or_404(EventPopup, event_id=event_id)
    event.delete()
    messages.success(request, "Event deleted successfully!")
    return redirect('view_event_history')



    



@user_passes_test(is_admin, login_url='/cafeteria_admin/admin_login/')
def manage_users(request):
    #Getting search query from GET parameters
    query = request.GET.get('q', '')  
    if query:
        #Searching users by username or email
        users = Profile.objects.filter(username__icontains=query) | \
                Profile.objects.filter(email__icontains=query)
    else:
        # If no search query, fetching all the users
        users = Profile.objects.all().order_by('username')

    return render(request, 'cafeteria_admin/manage_users.html', {'users': users, 'query': query})


@user_passes_test(is_admin, login_url='/cafeteria_admin/admin_login/')
def edit_user(request, user_id):
    #Fetching the user by ID or 404 if not found
    user = get_object_or_404(Profile, id=user_id)
    if request.method == 'POST':
        user.username = request.POST['username']
        user.phone = request.POST['phone']
        user.email = request.POST['email']
        user.save()
        messages.success(request, 'User updated successfully!')
        #Redirecting after successfully updating the user
        return redirect('manage_users')
    return render(request, 'cafeteria_admin/edit_user.html', {'user': user})

@user_passes_test(is_admin, login_url='/cafeteria_admin/admin_login/')
def delete_user(request, user_id):
    user = get_object_or_404(Profile, id=user_id)
    user.delete()
    messages.success(request, 'User deleted successfully!')
    return redirect('manage_users')

@user_passes_test(is_admin, login_url='/cafeteria_admin/admin_login/')
def update_user_password(request, user_id):
    user = get_object_or_404(User, id=user_id)
    
    if request.method == 'POST':
        new_password = request.POST['new_password']
        confirm_password = request.POST['confirm_password']

        if new_password != confirm_password:
            messages.error(request, "Passwords do not match.")
             # Return if passwords do not match
            return redirect('manage_users')

        #Updating the user's password
        user.password = make_password(new_password)
        user.save()
        messages.success(request, f"Password for {user.username} updated successfully.")
        #Redurecting after a updating password
        return redirect('manage_users')

    return render(request, 'cafeteria_admin/update_password.html', {'user': user})   



@user_passes_test(is_admin, login_url='/cafeteria_admin/admin_login/')
def manage_menu(request):
    #Getting search query from GET parameters
    query = request.GET.get('q', '')  
    if query:
        food_items = FoodItem.objects.filter(name__icontains=query) | FoodItem.objects.filter(category__icontains=query)
    else:
         # If no search query, fetch all food items
        food_items = FoodItem.objects.all().order_by('category')

    return render(request, 'cafeteria_admin/manage_menu.html', {'food_items': food_items, 'query': query})



@user_passes_test(is_admin, login_url='/cafeteria_admin/admin_login/')
def add_food_item(request):
    try:
        if request.method == 'POST':
            # Processing the form data
            form = FoodItemForm(request.POST, request.FILES)
            if form.is_valid():
                form.save()
                # Redirecting after adding the food item
                return redirect('manage_menu')
        else:
            # Creating an empty form for GET requests
            form = FoodItemForm()
        return render(request, 'cafeteria_admin/add_food.html', {'form': form})
    except Exception as e:
        logger.error(f"Error in add_food_item: {str(e)}")
        messages.error(request, "An error occurred while adding the food item. Please try again.")
        form = FoodItemForm(request.POST or None, request.FILES or None) if request.method == 'POST' else FoodItemForm()
        return render(request, 'cafeteria_admin/add_food.html', {'form': form})

@user_passes_test(is_admin, login_url='/cafeteria_admin/admin_login/')
def edit_food_item(request, food_id):
    food_item = get_object_or_404(FoodItem, id=food_id)
    if request.method == 'POST':
        #Binding form with food item instance
        form = FoodItemForm(request.POST, request.FILES, instance=food_item)
        if form.is_valid():
            form.save()
            messages.success(request, 'Food item updated successfully!')
            return redirect('manage_menu')
    else:
        #Pre-poluting rhe form with existing food item data
        form = FoodItemForm(instance=food_item)
    return render(request, 'cafeteria_admin/edit_food.html', {'form': form, 'food_item': food_item})

@user_passes_test(is_admin, login_url='/cafeteria_admin/admin_login/')
def delete_food_item(request, food_id):
    food_item = get_object_or_404(FoodItem, id=food_id)
    food_item.delete()
    messages.success(request, 'Food item deleted successfully!')
    return redirect('manage_menu')
  


#  Manage Orders
def manage_orders(request):
    #Getting search query from GET parameters
    query = request.GET.get('q')  

    if query:
        #Searching with the help of order id 
        orders = Order.objects.filter(id=query)  
    else:
        # Showing all the orders if no query is provided
        orders = Order.objects.all().order_by('-order_date')  

    return render(request, 'cafeteria_admin/manage_orders.html', {'orders': orders, 'query': query})

# Updating  Order Status
@user_passes_test(is_admin, login_url='/cafeteria_admin/admin_login/')
def update_order_status(request, order_id):
    if not request.user.is_staff:
        messages.error(request, "Access denied.")
        return redirect('manage_orders')

    order = get_object_or_404(Order, id=order_id)

    if request.method == "POST":
        new_status = request.POST.get("status")
        order.status = new_status
        order.save()
        messages.success(request, f"Order #{order_id} updated successfully!")
    
    return redirect('manage_orders')

# Deleting the Order
@user_passes_test(is_admin, login_url='/cafeteria_admin/admin_login/')
def delete_order(request, order_id):
    if not request.user.is_staff:
        messages.error(request, "Access denied.")
        return redirect('manage_orders')

    order = get_object_or_404(Order, id=order_id)
    order.delete()
    messages.success(request, f"Order #{order_id} deleted successfully!")
    
    return redirect('manage_orders')


# View Order Details
@user_passes_test(is_admin, login_url='/cafeteria_admin/admin_login/')
def manage_lost_found(request):
    query = request.GET.get('q', '')
    if query:
        items = LostFound.objects.filter(item_name__icontains=query) | LostFound.objects.filter(location__icontains=query)
    else:
        items = LostFound.objects.all()
    return render(request, 'cafeteria_admin/manage_lost_found.html', {'items': items, 'query': query})

@user_passes_test(is_admin, login_url='/cafeteria_admin/admin_login/')
def approve_lost_found(request, item_id):
    if request.method == 'POST':
        item = get_object_or_404(LostFound, id=item_id)
        item.status = 'approved'
        item.approved_by = request.user
        item.save()
        messages.success(request, f"Item '{item.item_name}' has been approved.")
    return redirect('manage_lost_found')

@user_passes_test(is_admin, login_url='/cafeteria_admin/admin_login/')
def resolve_lost_found(request, item_id):
    if request.method == 'POST':
        item = get_object_or_404(LostFound, id=item_id)
        item.status = 'resolved'
        item.save()
        messages.success(request, f"Item '{item.item_name}' has been marked as resolved.")
    return redirect('manage_lost_found')

@user_passes_test(is_admin, login_url='/cafeteria_admin/admin_login/')
def delete_lost_found(request, item_id):
    if request.method == 'POST':
        item = get_object_or_404(LostFound, id=item_id)
        item_name = item.item_name
        item.delete()
        messages.success(request, f"Lost item '{item_name}' deleted successfully.")
        return redirect('manage_lost_found')
    return redirect('manage_lost_found')

@user_passes_test(is_admin, login_url='/cafeteria_admin/admin_login/')
def manage_group_orders(request):
    query = request.GET.get('q', '')
    if query:
        group_orders = GroupOrder.objects.filter(code__icontains=query) | GroupOrder.objects.filter(leader__username__icontains=query)
    else:
        group_orders = GroupOrder.objects.all().order_by('-created_at')
    return render(request, 'cafeteria_admin/manage_group_orders.html', {'group_orders': group_orders, 'query': query})

@user_passes_test(is_admin, login_url='/cafeteria_admin/admin_login/')
def close_group_order(request, group_id):
    if request.method == 'POST':
        group = get_object_or_404(GroupOrder, id=group_id)
        if group.is_active:
            group.is_active = False
            order = Order.objects.create(
                user=group.leader,
                total_price=group.total_price,
                payment_method="Cash",
                remarks="Group Order Closed by Admin"
            )
            for item in group.group_items.all():
                OrderItem.objects.create(
                    order=order,
                    food_item=item.food_item,
                    quantity=item.quantity,
                    price=item.subtotal
                )
            group.save()
            messages.success(request, f"Group Order {group.code} closed and converted to Order #{order.id}.")
        else:
            messages.info(request, f"Group Order {group.code} is already closed.")
    return redirect('manage_group_orders')

@user_passes_test(is_admin, login_url='/cafeteria_admin/admin_login/')
def delete_group_order(request, group_id):
    if request.method == 'POST':
        group = get_object_or_404(GroupOrder, id=group_id)
        group.delete()
        messages.success(request, f"Group Order {group.code} deleted successfully!")
    return redirect('manage_group_orders')



@user_passes_test(is_admin, login_url='/cafeteria_admin/admin_login/')
def manage_feedback(request):
    try:
        query = request.GET.get('q', '')
        if query:
            feedbacks = Feedback.objects.filter(
                Q(content__icontains=query) | 
                Q(user__username__icontains=query) | 
                Q(replies__content__icontains=query)
            ).distinct().order_by('-created_at')
        else:
            feedbacks = Feedback.objects.all().order_by('-created_at')

        if request.method == "POST":
            action = request.POST.get('action')
            feedback_id = request.POST.get('feedback_id')
            feedback = get_object_or_404(Feedback, id=feedback_id)

            if action == "approve":
                feedback.is_approved = True
                feedback.save()
                messages.success(request, f"Feedback #{feedback_id} approved.")
            elif action == "delete":
                feedback.delete()
                messages.success(request, f"Feedback #{feedback_id} deleted.")
            return redirect('manage_feedback')

        return render(request, 'cafeteria_admin/manage_feedback.html', {'feedbacks': feedbacks, 'query': query})
    except Exception as e:
        logger.error(f"Error in manage_feedback: {str(e)}")
        messages.error(request, "An error occurred while managing feedback.")
        return redirect('cafeteria_admin_dashboard')

@user_passes_test(is_admin, login_url='/cafeteria_admin/admin_login/')
def admin_edit_reply(request, reply_id):
    reply = get_object_or_404(Reply, id=reply_id)
    if request.method == "POST":
        content = request.POST.get('content')
        if content:
            reply.content = content
            reply.save()
            messages.success(request, f"Reply #{reply_id} updated successfully!")
        else:
            messages.error(request, "Reply content cannot be empty.")
    return redirect('manage_feedback')

@user_passes_test(is_admin, login_url='/cafeteria_admin/admin_login/')
def admin_delete_reply(request, reply_id):
    reply = get_object_or_404(Reply, id=reply_id)
    if request.method == "POST":
        reply.delete()
        messages.success(request, f"Reply #{reply_id} deleted successfully!")
    return redirect('manage_feedback')




# Form for top-selling food filters
class TopSellingForm(forms.Form):
    food_name = forms.CharField(
        max_length=100,
        required=False,
        label='Food Name',
        widget=forms.TextInput(attrs={'placeholder': 'e.g., Momo', 'class': 'form-control'})
    )
    period = forms.ChoiceField(
        choices=[
            ('month', 'Last Month'),
            ('week', 'Last Week'),
            ('year', 'Last Year'),
            ('custom', 'Custom Range'),
        ],
        required=True,
        label='Time Period',
        widget=forms.Select(attrs={'class': 'form-control'})
    )
    start_date = forms.DateField(
        widget=forms.DateInput(attrs={'type': 'date', 'class': 'form-control'}),
        required=False,
        label='Start Date'
    )
    end_date = forms.DateField(
        widget=forms.DateInput(attrs={'type': 'date', 'class': 'form-control'}),
        required=False,
        label='End Date'
    )

    def clean(self):
        cleaned_data = super().clean()
        period = cleaned_data.get('period')
        start_date = cleaned_data.get('start_date')
        end_date = cleaned_data.get('end_date')
        if period == 'custom':
            if not (start_date and end_date):
                raise forms.ValidationError('Start and end dates are required for custom range.')
            if end_date < start_date:
                raise forms.ValidationError('End date must be after start date.')
        return cleaned_data

# Top-Selling Food Tracking
@user_passes_test(is_admin, login_url='/cafeteria_admin/admin_login/')
def top_selling_food(request):
    try:
        # Default: last 30 days
        default_end = timezone.now().date()
        default_start = default_end - timedelta(days=30)
        initial_data = {'period': 'month', 'start_date': default_start, 'end_date': default_end}
        form = TopSellingForm(request.GET or initial_data)

        top_items = []
        total_sales = 0
        selected_period = 'month'
        food_name = ''

        if form.is_valid():
            food_name = form.cleaned_data['food_name'].strip()
            period = form.cleaned_data['period']
            selected_period = period
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']

            # Set date range
            if period == 'week':
                start_date = default_end - timedelta(days=7)
                end_date = default_end
            elif period == 'month':
                start_date = default_end - timedelta(days=30)
                end_date = default_end
            elif period == 'year':
                start_date = default_end - timedelta(days=365)
                end_date = default_end

            # Convert to timezone-aware datetimes
            start_datetime = timezone.make_aware(datetime.combine(start_date, datetime.min.time()))
            end_datetime = timezone.make_aware(datetime.combine(end_date, datetime.max.time()))

            # Build query
            query = OrderItem.objects.filter(
                order__order_date__range=[start_datetime, end_datetime]
            )
            if food_name:
                query = query.filter(food_item__name__icontains=food_name)

            logger.debug(f"Querying OrderItem for {food_name or 'all items'}, {start_datetime} to {end_datetime}")
            top_items = query.values(
                'food_item__name', 'food_item__category'
            ).annotate(
                total_quantity=Sum('quantity'),
                total_revenue=Sum('price')
            ).order_by('-total_quantity')[:10]
            logger.debug(f"Top items: {list(top_items)}")
            total_sales = sum(item['total_revenue'] for item in top_items) or 0

        context = {
            'form': form,
            'top_items': top_items,
            'total_sales': total_sales,
            'selected_period': selected_period,
            'food_name': food_name,
            'start_date': start_date if form.is_valid() else default_start,
            'end_date': end_date if form.is_valid() else default_end,
        }
        return render(request, 'cafeteria_admin/top_selling_food.html', context)
    except Exception as e:
        logger.error(f"Error in top_selling_food: {str(e)}")
        messages.error(request, "An error occurred while loading top-selling items.")
        return render(request, 'cafeteria_admin/top_selling_food.html', {'form': form})
        


# Manage Payments
@user_passes_test(is_admin, login_url='/cafeteria_admin/admin_login/')
def manage_payments(request):
    query = request.GET.get('q', '')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    payment_method = request.GET.get('payment_method', '')

    orders = Order.objects.select_related('user')

    # Applying the  date range filter
    if start_date and end_date:
        try:
            start = timezone.make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
            end = timezone.make_aware(datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1))
            orders = orders.filter(order_date__range=[start, end])
        except ValueError:
            messages.error(request, "Invalid date format.")
    else:
        # Defaulting it to last 30 days
        default_start = timezone.now() - timedelta(days=30)
        orders = orders.filter(order_date__gte=default_start)
        start_date = default_start.strftime('%Y-%m-%d')
        end_date = timezone.now().strftime('%Y-%m-%d')

    # Applying the payment method filter
    if payment_method in ['Cash', 'Online']:
        orders = orders.filter(payment_method=payment_method)

    # Applying the text search
    if query:
        orders = orders.filter(Q(id__icontains=query) | Q(user__username__icontains=query))

    orders = orders.order_by('-order_date')

    # Calculate total amount for filtered orders
    total_amount = orders.aggregate(total=Sum('total_price'))['total'] or 0

    logger.debug(f"manage_payments: query='{query}', payment_method='{payment_method}', "
                 f"start_date='{start_date}', end_date='{end_date}', "
                 f"orders count={orders.count()}, total_amount={total_amount}")

    context = {
        'orders': orders,
        'query': query,
        'start_date': start_date,
        'end_date': end_date,
        'payment_method': payment_method,
        'total_amount': float(total_amount),
    }
    return render(request, 'cafeteria_admin/manage_payments.html', context)

# Exporting the Payments to Excel
@user_passes_test(is_admin, login_url='/cafeteria_admin/admin_login/')
def export_payments_to_excel(request):
    query = request.GET.get('q', '')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    payment_method = request.GET.get('payment_method', '')

    orders = Order.objects.select_related('user')

    # Applying  the  date range filter
    if start_date and end_date:
        try:
            start = timezone.make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
            end = timezone.make_aware(datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1))
            orders = orders.filter(order_date__range=[start, end])
        except ValueError:
            messages.error(request, "Invalid date format.")
            return redirect('manage_payments')
    else:
        default_start = timezone.now() - timedelta(days=30)
        orders = orders.filter(order_date__gte=default_start)

    # Applying the payment method filter
    if payment_method in ['Cash', 'Online']:
        orders = orders.filter(payment_method=payment_method)

    # Applying the text search
    if query:
        orders = orders.filter(Q(id__icontains=query) | Q(user__username__icontains=query))

    orders = orders.order_by('-order_date')

    # Calculating the total amount
    total_amount = orders.aggregate(total=Sum('total_price'))['total'] or 0

    # Creating the Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"

    # Defining the headers
    headers = ["Order ID", "Username", "Amount (Rs.)", "Payment Method", "Date"]
    ws.append(headers)

    # Styling the headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for order in orders:
        ws.append([
            order.id,
            order.user.username,
            float(order.total_price),
            order.get_payment_method_display(),
            order.order_date.strftime('%b %d, %Y')
        ])


    total_row = ["", "", f"Total Amount: Rs. {float(total_amount):.2f}", "", ""]
    ws.append(total_row)
    total_cell = ws.cell(row=ws.max_row, column=3)
    total_cell.font = Font(bold=True, color="99180D")
    total_cell.alignment = Alignment(horizontal='left')


    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width


    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        content=output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=payments_export.xlsx'
    return response